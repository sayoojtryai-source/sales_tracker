from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation

from flask import (
    Flask,
    abort,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import (
    LoginManager,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from openpyxl import Workbook
from sqlalchemy import func, text

from config import Config
from models import Customer, Order, OrderItem, Product, User, db


def create_app(config_class: type = Config) -> Flask:
    app = Flask(__name__)
    app.config.from_object(config_class)

    db.init_app(app)

    login_manager = LoginManager()
    login_manager.login_view = "login"
    login_manager.login_message_category = "info"
    login_manager.init_app(app)

    @login_manager.user_loader
    def load_user(user_id: str):
        return db.session.get(User, int(user_id))

    @app.context_processor
    def inject_branding():
        return {
            "brand_name": app.config["BRAND_NAME"],
            "brand_tagline": app.config["BRAND_TAGLINE"],
            "currency": app.config["CURRENCY_SYMBOL"],
        }

    register_routes(app)

    with app.app_context():
        db.create_all()
        _run_migrations()
        _ensure_default_admin()

    return app


def _run_migrations() -> None:
    with db.engine.connect() as conn:
        result = conn.execute(text("PRAGMA table_info(orders)"))
        cols = [row[1] for row in result]
        if "customer_id" not in cols:
            conn.execute(text(
                "ALTER TABLE orders ADD COLUMN customer_id INTEGER REFERENCES customers(id)"
            ))
            conn.commit()


def _ensure_default_admin() -> None:
    if User.query.count() == 0:
        admin = User(username="admin", is_admin=True)
        admin.set_password("admin123")
        db.session.add(admin)
        db.session.commit()


def _parse_decimal(value: str) -> Decimal:
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        raise ValueError(f"Invalid number: {value}")


def _parse_int(value: str) -> int:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        raise ValueError(f"Invalid integer: {value}")


def _day_bounds(target: date) -> tuple[datetime, datetime]:
    start = datetime.combine(target, datetime.min.time())
    end = start + timedelta(days=1)
    return start, end


def register_routes(app: Flask) -> None:
    # ----- Auth -----
    @app.route("/login", methods=["GET", "POST"])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for("dashboard"))
        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            user = User.query.filter_by(username=username).first()
            if user and user.check_password(password):
                login_user(user)
                return redirect(url_for("dashboard"))
            flash("Invalid username or password.", "danger")
        return render_template("login.html")

    @app.route("/logout")
    @login_required
    def logout():
        logout_user()
        return redirect(url_for("login"))

    # ----- Dashboard -----
    @app.route("/")
    @login_required
    def dashboard():
        today = date.today()
        today_start, today_end = _day_bounds(today)

        # Order counts
        total_orders_all = db.session.query(func.count(Order.id)).scalar() or 0
        total_orders_today = (
            db.session.query(func.count(Order.id))
            .filter(Order.created_at >= today_start, Order.created_at < today_end)
            .scalar()
            or 0
        )
        total_orders_past = total_orders_all - total_orders_today

        # Sales totals
        def _sum_between(start, end):
            return (
                db.session.query(func.coalesce(func.sum(Order.total_amount), 0))
                .filter(Order.created_at >= start, Order.created_at < end)
                .scalar()
                or 0
            )

        sales_today = float(_sum_between(today_start, today_end))
        sales_all = float(
            db.session.query(func.coalesce(func.sum(Order.total_amount), 0)).scalar()
            or 0
        )
        sales_past = sales_all - sales_today

        products = Product.query.order_by(Product.name.asc()).all()
        total_stock = sum(p.quantity for p in products)
        low_stock = [p for p in products if p.is_active and p.quantity <= 5]

        recent_orders = (
            Order.query.order_by(Order.created_at.desc()).limit(5).all()
        )

        return render_template(
            "dashboard.html",
            total_orders_all=total_orders_all,
            total_orders_today=total_orders_today,
            total_orders_past=total_orders_past,
            sales_today=sales_today,
            sales_past=sales_past,
            sales_all=sales_all,
            products=products,
            total_stock=total_stock,
            low_stock=low_stock,
            recent_orders=recent_orders,
        )

    # ----- Products -----
    @app.route("/products")
    @login_required
    def products_list():
        products = Product.query.order_by(Product.name.asc()).all()
        return render_template("products.html", products=products)

    @app.route("/products/new", methods=["GET", "POST"])
    @login_required
    def product_create():
        if request.method == "POST":
            try:
                name = request.form.get("name", "").strip()
                if not name:
                    raise ValueError("Name is required.")
                product = Product(
                    name=name,
                    category=request.form.get("category", "Other").strip() or "Other",
                    price=_parse_decimal(request.form.get("price", "0")),
                    quantity=_parse_int(request.form.get("quantity", "0")),
                    description=request.form.get("description", "").strip(),
                    is_active=bool(request.form.get("is_active")),
                )
                db.session.add(product)
                db.session.commit()
                flash(f"Product '{product.name}' added.", "success")
                return redirect(url_for("products_list"))
            except ValueError as exc:
                flash(str(exc), "danger")
            except Exception as exc:  # pragma: no cover
                db.session.rollback()
                flash(f"Could not save product: {exc}", "danger")
        return render_template("product_form.html", product=None)

    @app.route("/products/<int:product_id>/edit", methods=["GET", "POST"])
    @login_required
    def product_edit(product_id: int):
        product = db.session.get(Product, product_id) or abort(404)
        if request.method == "POST":
            try:
                product.name = request.form.get("name", product.name).strip()
                product.category = (
                    request.form.get("category", product.category).strip() or "Other"
                )
                product.price = _parse_decimal(
                    request.form.get("price", str(product.price))
                )
                product.quantity = _parse_int(
                    request.form.get("quantity", str(product.quantity))
                )
                product.description = request.form.get(
                    "description", product.description or ""
                ).strip()
                product.is_active = bool(request.form.get("is_active"))
                db.session.commit()
                flash(f"Product '{product.name}' updated.", "success")
                return redirect(url_for("products_list"))
            except ValueError as exc:
                flash(str(exc), "danger")
            except Exception as exc:  # pragma: no cover
                db.session.rollback()
                flash(f"Could not update product: {exc}", "danger")
        return render_template("product_form.html", product=product)

    @app.route("/products/<int:product_id>/delete", methods=["POST"])
    @login_required
    def product_delete(product_id: int):
        product = db.session.get(Product, product_id) or abort(404)
        if product.order_items:
            product.is_active = False
            db.session.commit()
            flash(
                f"'{product.name}' is referenced by past orders and was deactivated instead of deleted.",
                "warning",
            )
        else:
            db.session.delete(product)
            db.session.commit()
            flash(f"Product '{product.name}' deleted.", "success")
        return redirect(url_for("products_list"))

    # ----- Orders -----
    @app.route("/orders")
    @login_required
    def orders_list():
        scope = request.args.get("scope", "all")  # all | today | past
        query = Order.query
        today_start, today_end = _day_bounds(date.today())
        if scope == "today":
            query = query.filter(
                Order.created_at >= today_start, Order.created_at < today_end
            )
        elif scope == "past":
            query = query.filter(Order.created_at < today_start)
        orders = query.order_by(Order.created_at.desc()).all()
        return render_template("orders.html", orders=orders, scope=scope)

    @app.route("/orders/new", methods=["GET", "POST"])
    @login_required
    def order_create():
        products = (
            Product.query.filter_by(is_active=True)
            .order_by(Product.name.asc())
            .all()
        )
        if request.method == "POST":
            try:
                customer_name = (
                    request.form.get("customer_name", "").strip() or "Walk-in"
                )
                customer_phone = request.form.get("customer_phone", "").strip()
                notes = request.form.get("notes", "").strip()

                product_ids = request.form.getlist("product_id")
                quantities = request.form.getlist("quantity")

                items: list[tuple[Product, int]] = []
                for pid_raw, qty_raw in zip(product_ids, quantities):
                    if not pid_raw:
                        continue
                    qty = _parse_int(qty_raw or "0")
                    if qty <= 0:
                        continue
                    product = db.session.get(Product, int(pid_raw))
                    if not product:
                        raise ValueError(f"Unknown product id: {pid_raw}")
                    if qty > product.quantity:
                        raise ValueError(
                            f"Not enough stock for '{product.name}' "
                            f"(have {product.quantity}, need {qty})."
                        )
                    items.append((product, qty))

                if not items:
                    raise ValueError("Add at least one item to the order.")

                order = Order(
                    customer_name=customer_name,
                    customer_phone=customer_phone,
                    notes=notes,
                    total_amount=Decimal("0"),
                )
                total = Decimal("0")
                for product, qty in items:
                    line_total = (product.price or Decimal("0")) * qty
                    total += line_total
                    order.items.append(
                        OrderItem(
                            product_id=product.id,
                            product_name=product.name,
                            unit_price=product.price,
                            quantity=qty,
                            line_total=line_total,
                        )
                    )
                    product.quantity -= qty  # decrement stock
                order.total_amount = total
                db.session.add(order)
                if customer_phone:
                    cust = Customer.query.filter_by(phone=customer_phone).first()
                    if cust is None:
                        cust = Customer(name=customer_name, phone=customer_phone)
                        db.session.add(cust)
                    elif customer_name != "Walk-in":
                        cust.name = customer_name
                    order.customer = cust
                db.session.commit()
                flash(
                    f"Order #{order.id} created. Total {app.config['CURRENCY_SYMBOL']}{total:.2f}.",
                    "success",
                )
                return redirect(url_for("order_detail", order_id=order.id))
            except ValueError as exc:
                db.session.rollback()
                flash(str(exc), "danger")
            except Exception as exc:  # pragma: no cover
                db.session.rollback()
                flash(f"Could not create order: {exc}", "danger")
        return render_template("order_form.html", products=products)

    @app.route("/orders/<int:order_id>")
    @login_required
    def order_detail(order_id: int):
        order = db.session.get(Order, order_id) or abort(404)
        return render_template("order_detail.html", order=order)

    @app.route("/orders/<int:order_id>/edit", methods=["GET", "POST"])
    @login_required
    def order_edit(order_id: int):
        order = db.session.get(Order, order_id) or abort(404)
        if request.method == "POST":
            try:
                order.customer_name = request.form.get("customer_name", "").strip() or "Walk-in"
                order.customer_phone = request.form.get("customer_phone", "").strip()
                order.notes = request.form.get("notes", "").strip()
                db.session.commit()
                flash(f"Order #{order.id} updated.", "success")
                return redirect(url_for("order_detail", order_id=order.id))
            except Exception as exc:
                db.session.rollback()
                flash(f"Could not update order: {exc}", "danger")
        return render_template("order_edit.html", order=order)

    @app.route("/orders/<int:order_id>/delete", methods=["POST"])
    @login_required
    def order_delete(order_id: int):
        order = db.session.get(Order, order_id) or abort(404)
        restored = 0
        for item in order.items:
            if item.product is not None:
                item.product.quantity += item.quantity
                restored += item.quantity
        db.session.delete(order)
        db.session.commit()
        flash(
            f"Order #{order_id} deleted. {restored} unit(s) returned to stock.",
            "success",
        )
        return redirect(url_for("orders_list"))

    # ----- Customers -----
    @app.route("/customers")
    @login_required
    def customers_list():
        rows = (
            db.session.query(
                Customer,
                func.count(Order.id).label("order_count"),
                func.coalesce(func.sum(Order.total_amount), 0).label("total_spent"),
            )
            .outerjoin(Order, Order.customer_id == Customer.id)
            .group_by(Customer.id)
            .order_by(Customer.name.asc())
            .all()
        )
        return render_template("customers.html", rows=rows)

    @app.route("/customers/new", methods=["GET", "POST"])
    @login_required
    def customer_create():
        if request.method == "POST":
            try:
                name = request.form.get("name", "").strip()
                phone = request.form.get("phone", "").strip()
                if not name:
                    raise ValueError("Name is required.")
                if not phone:
                    raise ValueError("Phone is required.")
                if Customer.query.filter_by(phone=phone).first():
                    raise ValueError(f"A customer with phone {phone} already exists.")
                cust = Customer(
                    name=name,
                    phone=phone,
                    email=request.form.get("email", "").strip(),
                    address=request.form.get("address", "").strip(),
                    notes=request.form.get("notes", "").strip(),
                )
                db.session.add(cust)
                db.session.commit()
                flash(f"Customer '{name}' added.", "success")
                return redirect(url_for("customers_list"))
            except ValueError as exc:
                db.session.rollback()
                flash(str(exc), "danger")
        return render_template("customer_form.html", customer=None)

    @app.route("/customers/<int:customer_id>")
    @login_required
    def customer_detail(customer_id: int):
        cust = db.session.get(Customer, customer_id) or abort(404)
        orders = (
            Order.query.filter_by(customer_id=customer_id)
            .order_by(Order.created_at.desc())
            .all()
        )
        total_spent = sum(o.total_float for o in orders)
        return render_template(
            "customer_detail.html", customer=cust, orders=orders, total_spent=total_spent
        )

    @app.route("/export/customers.xlsx")
    @login_required
    def export_customers():
        rows = (
            db.session.query(
                Customer,
                func.count(Order.id).label("order_count"),
                func.coalesce(func.sum(Order.total_amount), 0).label("total_spent"),
            )
            .outerjoin(Order, Order.customer_id == Customer.id)
            .group_by(Customer.id)
            .order_by(Customer.name.asc())
            .all()
        )
        wb = Workbook()
        ws = wb.active
        ws.title = "Customers"
        ws.append(["ID", "Name", "Phone", "Email", "Address", "Notes", "Joined", "Total Orders", "Total Spent (INR)"])
        for cust, order_count, total_spent in rows:
            ws.append([
                cust.id,
                cust.name,
                cust.phone,
                cust.email or "",
                cust.address or "",
                cust.notes or "",
                cust.created_at.strftime("%Y-%m-%d") if cust.created_at else "",
                order_count,
                float(total_spent or 0),
            ])
        return _send_workbook(wb, f"caek-customers-{date.today().isoformat()}.xlsx")

    # ----- Exports -----
    @app.route("/export/products.xlsx")
    @login_required
    def export_products():
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(
            [
                "ID",
                "Name",
                "Category",
                "Price (INR)",
                "Quantity",
                "Active",
                "Description",
                "Created",
                "Updated",
            ]
        )
        for p in Product.query.order_by(Product.name.asc()).all():
            ws.append(
                [
                    p.id,
                    p.name,
                    p.category,
                    float(p.price or 0),
                    p.quantity,
                    "Yes" if p.is_active else "No",
                    p.description or "",
                    p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else "",
                    p.updated_at.strftime("%Y-%m-%d %H:%M") if p.updated_at else "",
                ]
            )
        return _send_workbook(wb, f"caek-products-{date.today().isoformat()}.xlsx")

    @app.route("/export/sales.xlsx")
    @login_required
    def export_sales():
        scope = request.args.get("scope", "all")  # all | today | past
        query = Order.query
        today_start, today_end = _day_bounds(date.today())
        if scope == "today":
            query = query.filter(
                Order.created_at >= today_start, Order.created_at < today_end
            )
        elif scope == "past":
            query = query.filter(Order.created_at < today_start)
        orders = query.order_by(Order.created_at.desc()).all()

        wb = Workbook()
        summary = wb.active
        summary.title = "Orders"
        summary.append(
            [
                "Order ID",
                "Date",
                "Time",
                "Customer",
                "Phone",
                "Items",
                "Total (INR)",
                "Notes",
            ]
        )
        for o in orders:
            summary.append(
                [
                    o.id,
                    o.created_at.strftime("%Y-%m-%d"),
                    o.created_at.strftime("%H:%M"),
                    o.customer_name,
                    o.customer_phone or "",
                    sum(i.quantity for i in o.items),
                    float(o.total_amount or 0),
                    o.notes or "",
                ]
            )

        lines = wb.create_sheet("Line Items")
        lines.append(
            [
                "Order ID",
                "Date",
                "Product",
                "Unit Price (INR)",
                "Quantity",
                "Line Total (INR)",
            ]
        )
        for o in orders:
            for item in o.items:
                lines.append(
                    [
                        o.id,
                        o.created_at.strftime("%Y-%m-%d"),
                        item.product_name,
                        float(item.unit_price or 0),
                        item.quantity,
                        float(item.line_total or 0),
                    ]
                )

        return _send_workbook(
            wb, f"caek-sales-{scope}-{date.today().isoformat()}.xlsx"
        )


def _send_workbook(wb: Workbook, filename: str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


app = create_app()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)
